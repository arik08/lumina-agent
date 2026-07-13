from __future__ import annotations

from io import BytesIO
from typing import cast

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from .model import ReportDocument
from .theme import COBALT_HEX, INK_HEX, LIGHT_BLUE_HEX


_COBALT = COBALT_HEX
_LIGHT_BLUE = LIGHT_BLUE_HEX
_INK = INK_HEX
_WHITE = "FFFFFF"
_THIN = Side(style="thin", color="DCE3EF")


def generate_xlsx(report: ReportDocument) -> bytes:
    workbook = Workbook()
    sheet = cast(Worksheet, workbook.active)
    sheet.title = "보고서"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A5"
    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 24
    sheet.column_dimensions["C"].width = 48
    sheet.column_dimensions["D"].width = 24

    sheet.merge_cells("A1:D1")
    _set_text(sheet, 1, 1, report.title)
    sheet["A1"].font = Font(name="맑은 고딕", size=20, bold=True, color=_WHITE)
    sheet["A1"].fill = PatternFill("solid", fgColor=_COBALT)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 34

    sheet.merge_cells("A2:D3")
    _set_text(sheet, 2, 1, report.executive_summary)
    sheet["A2"].font = Font(name="맑은 고딕", size=11, color=_INK)
    sheet["A2"].fill = PatternFill("solid", fgColor="F5F7FA")
    sheet["A2"].alignment = Alignment(wrap_text=True, vertical="center")

    row = 5
    row = _section_header(sheet, row, "핵심 지표")
    _set_text(sheet, row, 1, "지표")
    _set_text(sheet, row, 2, "값")
    _style_table_header(sheet, row, 1, 2)
    row += 1
    for metric in report.metrics:
        _set_text(sheet, row, 1, metric.label)
        _set_text(sheet, row, 2, metric.value)
        _style_table_row(sheet, row, 1, 2)
        row += 1

    row += 1
    row = _section_header(sheet, row, "원 요청")
    sheet.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=4)
    _set_text(sheet, row, 1, report.request)
    sheet.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="top")
    sheet.cell(row, 1).fill = PatternFill("solid", fgColor="F5F7FA")
    sheet.row_dimensions[row].height = 42
    row += 3

    row = _section_header(sheet, row, "상세 결과")
    detail_header_row = row
    for values in (("구분", "제목", "내용", "세부 항목"),):
        for column, value in enumerate(values, start=1):
            _set_text(sheet, row, column, value)
    _style_table_header(sheet, row, 1, 4)
    row += 1
    for index, section in enumerate(report.sections, start=1):
        _set_text(sheet, row, 1, f"섹션 {index}")
        _set_text(sheet, row, 2, section.heading)
        _set_text(sheet, row, 3, section.body)
        _set_text(sheet, row, 4, "\n".join(f"• {item}" for item in section.bullets))
        _style_table_row(sheet, row, 1, 4)
        sheet.row_dimensions[row].height = max(32, 15 * (len(section.bullets) + 1))
        row += 1
    sheet.auto_filter.ref = f"A{detail_header_row}:D{row - 1}"

    if report.action_items:
        row += 1
        row = _section_header(sheet, row, "후속 조치")
        _set_text(sheet, row, 1, "번호")
        _set_text(sheet, row, 2, "조치 항목")
        _style_table_header(sheet, row, 1, 4)
        row += 1
        for index, action in enumerate(report.action_items, start=1):
            sheet.cell(row, 1, index)
            sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            _set_text(sheet, row, 2, action)
            _style_table_row(sheet, row, 1, 4)
            row += 1

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _section_header(sheet: Worksheet, row: int, title: str) -> int:
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = _set_text(sheet, row, 1, title)
    cell.font = Font(name="맑은 고딕", size=12, bold=True, color=_COBALT)
    cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[row].height = 24
    return row + 1


def _style_table_header(sheet: Worksheet, row: int, start: int, end: int) -> None:
    for column in range(start, end + 1):
        cell = sheet.cell(row, column)
        cell.font = Font(name="맑은 고딕", bold=True, color=_INK)
        cell.fill = PatternFill("solid", fgColor=_LIGHT_BLUE)
        cell.border = Border(top=_THIN, bottom=_THIN)
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def _style_table_row(sheet: Worksheet, row: int, start: int, end: int) -> None:
    for column in range(start, end + 1):
        cell = sheet.cell(row, column)
        cell.font = Font(name="맑은 고딕", size=10, color=_INK)
        cell.border = Border(bottom=_THIN)
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def _set_text(sheet: Worksheet, row: int, column: int, value: str) -> Cell:
    cell = cast(Cell, sheet.cell(row, column))
    cell.value = value
    cell.data_type = "s"
    return cell
