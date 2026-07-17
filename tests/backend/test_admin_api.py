from __future__ import annotations

import json
from datetime import timedelta
from io import BytesIO
from pathlib import Path

from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select

from lumina.api.routes import admin
from lumina.config import Settings
from lumina.db import SessionLocal
from lumina.main import create_app
from lumina.models import (
    Conversation,
    AuditEvent,
    Message,
    Organization,
    Project,
    QueuedMessage,
    Run,
    User,
    utc_now,
)


def _test_app(tmp_path: Path) -> FastAPI:
    settings = Settings(
        environment="test",
        DATABASE_URL=f"sqlite:///{(tmp_path / 'admin.db').as_posix()}",
        data_dir=tmp_path,
        files_dir=tmp_path / "files",
        artifacts_dir=tmp_path / "artifacts",
        cookie_secure=False,
    )
    app = create_app(settings)
    static_mounts = [
        route for route in app.routes if getattr(route, "name", None) == "web"
    ]
    for mount in static_mounts:
        app.router.routes.remove(mount)
    if not any(
        getattr(route, "original_router", None) is admin.router for route in app.routes
    ):
        app.include_router(admin.router, prefix="/api")
    app.router.routes.extend(static_mounts)
    return app


def _login(client: TestClient, login_name: str, password: str) -> str:
    response = client.post(
        "/api/auth/login",
        json={
            "loginName": login_name,
            "loginDomain": "posco.com",
            "password": password,
        },
    )
    assert response.status_code == 200, response.text
    return response.json()["csrfToken"]


def _create_user(
    client: TestClient,
    csrf: str,
    *,
    login_name: str,
    password: str = "initial-password",
    role: str = "user",
) -> dict[str, object]:
    response = client.post(
        "/api/admin/users",
        headers={"X-CSRF-Token": csrf},
        json={
            "loginName": login_name,
            "loginDomain": "posco.com",
            "password": password,
            "displayName": login_name.title(),
            "affiliation": "AI 플랫폼팀",
            "role": role,
            "status": "active",
            "mustChangePassword": False,
        },
    )
    assert response.status_code == 201, response.text
    return response.json()


def test_admin_run_safety_settings_and_emergency_stop(tmp_path: Path) -> None:
    app = _test_app(tmp_path)
    with TestClient(app) as admin_client:
        csrf = _login(admin_client, "admin", "1")

        defaults = admin_client.get("/api/admin/run-safety")
        assert defaults.status_code == 200, defaults.text
        assert defaults.json() == {
            "maxModelTurns": 400,
            "maxTotalTokens": 4_000_000,
            "maxElapsedMinutes": 10_080,
            "maxCostUsd": 100.0,
        }
        assert (
            admin_client.patch(
                "/api/admin/run-safety",
                json={
                    "maxModelTurns": 300,
                    "maxTotalTokens": 3_000_000,
                    "maxElapsedMinutes": 480,
                    "maxCostUsd": 75,
                },
            ).status_code
            == 403
        )

        updated = admin_client.patch(
            "/api/admin/run-safety",
            headers={"X-CSRF-Token": csrf},
            json={
                "maxModelTurns": 300,
                "maxTotalTokens": 3_000_000,
                "maxElapsedMinutes": 480,
                "maxCostUsd": 75,
            },
        )
        assert updated.status_code == 200, updated.text
        assert updated.json()["maxTotalTokens"] == 3_000_000

        with SessionLocal() as db:
            admin_user = db.scalar(
                select(User).where(User.login_id == "admin@posco.com")
            )
            assert admin_user is not None
            organization = db.get(Organization, admin_user.organization_id)
            project = db.scalar(
                select(Project).where(Project.owner_user_id == admin_user.id)
            )
            assert organization is not None and project is not None
            conversation = Conversation(
                organization_id=organization.id,
                project_id=project.id,
                owner_user_id=admin_user.id,
                title="비상 중단 테스트",
            )
            db.add(conversation)
            db.flush()
            assert organization.run_safety_settings_json["max_model_turns"] == 300
            active_run = Run(
                organization_id=organization.id,
                project_id=conversation.project_id,
                conversation_id=conversation.id,
                user_id=admin_user.id,
                status="model_streaming",
                provider_id="mock",
                model_key="mock-agent",
                runtime_model_id="mock-agent",
                model_display_name="Mock Agent",
                snapshot_json={"limits": {}},
                usage_json={},
                idempotency_key="admin-emergency-active",
            )
            queued_run = Run(
                organization_id=organization.id,
                project_id=conversation.project_id,
                conversation_id=conversation.id,
                user_id=admin_user.id,
                status="queued",
                provider_id="mock",
                model_key="mock-agent",
                runtime_model_id="mock-agent",
                model_display_name="Mock Agent",
                snapshot_json={"limits": {}},
                usage_json={},
                idempotency_key="admin-emergency-queued",
            )
            db.add_all((active_run, queued_run))
            db.flush()
            queued_message = QueuedMessage(
                conversation_id=conversation.id,
                user_id=admin_user.id,
                position=1,
                message_text="비상 중단할 다음 요청",
                idempotency_key="admin-emergency-message",
            )
            db.add(queued_message)
            db.commit()
            run_ids = {active_run.id, queued_run.id}
            queued_message_id = queued_message.id

        stopped = admin_client.post(
            "/api/admin/run-safety/emergency-stop",
            headers={"X-CSRF-Token": csrf},
            json={"reason": "테스트 비상 중단"},
        )
        assert stopped.status_code == 200, stopped.text
        assert stopped.json()["cancelledRunCount"] == 2
        assert stopped.json()["cancelledQueuedMessageCount"] == 1

        with SessionLocal() as db:
            statuses = set(db.scalars(select(Run.status).where(Run.id.in_(run_ids))))
            assert statuses == {"cancelled"}
            queued_message = db.get(QueuedMessage, queued_message_id)
            assert queued_message is not None
            assert queued_message.status == "cancelled"
            assert queued_message.cancelled_at is not None

        audit = admin_client.get("/api/admin/audit-events?action=admin_all_runs_killed")
        assert audit.status_code == 200
        assert audit.json()["items"][0]["metadata"]["cancelled_run_count"] == 2


def test_admin_user_lifecycle_permissions_and_audit(tmp_path: Path) -> None:
    app = _test_app(tmp_path)
    with TestClient(app) as admin_client:
        admin_csrf = _login(admin_client, "admin", "1")
        assert (
            admin_client.post(
                "/api/admin/users",
                json={
                    "loginName": "csrf-probe",
                    "loginDomain": "posco.com",
                    "password": "not-created",
                },
            ).status_code
            == 403
        )
        user = _create_user(admin_client, admin_csrf, login_name="operator")
        user_id = str(user["id"])

        listed = admin_client.get("/api/admin/users?query=OPER ATOR")
        assert listed.status_code == 200
        assert listed.json()["items"] == []
        listed = admin_client.get("/api/admin/users?query=OPER")
        assert listed.status_code == 200
        assert [item["loginId"] for item in listed.json()["items"]] == [
            "operator@posco.com"
        ]
        assert listed.json()["items"][0]["affiliation"] == "AI 플랫폼팀"
        assert "passwordHash" not in listed.text

        user_client = TestClient(app)
        try:
            _login(user_client, "operator", "initial-password")
            forbidden = user_client.get("/api/admin/users")
            assert forbidden.status_code == 403

            disabled = admin_client.patch(
                f"/api/admin/users/{user_id}",
                headers={"X-CSRF-Token": admin_csrf},
                json={"status": "disabled"},
            )
            assert disabled.status_code == 200, disabled.text
            assert disabled.json()["status"] == "disabled"
            assert user_client.get("/api/auth/session").status_code == 401

            enabled = admin_client.patch(
                f"/api/admin/users/{user_id}",
                headers={"X-CSRF-Token": admin_csrf},
                json={"status": "active"},
            )
            assert enabled.status_code == 200

            reset = admin_client.post(
                f"/api/admin/users/{user_id}/reset-password",
                headers={"X-CSRF-Token": admin_csrf},
                json={
                    "newPassword": "replacement-password",
                    "mustChangePassword": True,
                },
            )
            assert reset.status_code == 200, reset.text
            assert reset.json()["user"]["mustChangePassword"] is True
            assert "replacement-password" not in reset.text
            assert (
                user_client.post(
                    "/api/auth/login",
                    json={
                        "loginName": "operator",
                        "loginDomain": "posco.com",
                        "password": "initial-password",
                    },
                ).status_code
                == 401
            )
            _login(user_client, "operator", "replacement-password")
        finally:
            user_client.close()

        admin_user = next(
            item
            for item in admin_client.get("/api/admin/users").json()["items"]
            if item["loginId"] == "admin@posco.com"
        )
        last_admin = admin_client.patch(
            f"/api/admin/users/{admin_user['id']}",
            headers={"X-CSRF-Token": admin_csrf},
            json={"status": "disabled"},
        )
        assert last_admin.status_code == 409
        assert last_admin.json()["code"] == "last_active_admin"

        audit = admin_client.get("/api/admin/audit-events?target_id=" + user_id)
        assert audit.status_code == 200
        actions = {item["action"] for item in audit.json()["items"]}
        assert {"user_created", "user_disabled", "password_reset_issued"} <= actions
        actor_by_action = {
            item["action"]: item["actorLoginId"] for item in audit.json()["items"]
        }
        assert actor_by_action["user_created"] == "admin@posco.com"
        assert actor_by_action["password_reset_issued"] == "admin@posco.com"
        assert "operator@posco.com" in actor_by_action.values()
        assert "replacement-password" not in audit.text
        assert "passwordHash" not in audit.text


def test_admin_usage_statistics_are_organization_scoped_and_admin_only(
    tmp_path: Path,
) -> None:
    app = _test_app(tmp_path)
    with TestClient(app) as admin_client:
        admin_csrf = _login(admin_client, "admin", "1")
        _create_user(admin_client, admin_csrf, login_name="analyst")
        user_client = TestClient(app)
        try:
            _login(user_client, "analyst", "initial-password")
            assert user_client.get("/api/admin/usage-statistics").status_code == 403
        finally:
            user_client.close()

        response = admin_client.get("/api/admin/usage-statistics?days=30")
        assert response.status_code == 200, response.text
        payload = response.json()
        assert payload["timezone"] == "Asia/Seoul"
        assert payload["periodDays"] == 30
        assert len(payload["trend"]) == 30
        assert payload["summary"]["dau"] >= 2
        assert payload["summary"]["mau"] >= 2
        assert payload["summary"]["stickinessPercent"] >= 0
        analyst = next(
            item for item in payload["users"] if item["loginId"] == "analyst@posco.com"
        )
        assert analyst["loginCount"] == 1
        assert analyst["activeDays"] == 1
        assert analyst["inactiveDays"] == 0
        assert analyst["inputTokens"] == 0
        assert analyst["cachedInputTokens"] == 0
        assert analyst["cacheHitRatioPercent"] == 0
        assert analyst["outputTokens"] == 0

        audit = admin_client.get(
            "/api/admin/audit-events?action=admin_usage_statistics_viewed"
        )
        assert audit.status_code == 200
        assert audit.json()["items"][0]["metadata"]["days"] == 30

        all_time = admin_client.get("/api/admin/usage-statistics?days=0")
        assert all_time.status_code == 200, all_time.text
        assert all_time.json()["periodDays"] >= 1
        assert len(all_time.json()["trend"]) == all_time.json()["periodDays"]
        assert all_time.json()["summary"]["runs"] >= payload["summary"]["runs"]


def test_admin_audit_traffic_returns_complete_minute_buckets(tmp_path: Path) -> None:
    app = _test_app(tmp_path)
    with TestClient(app) as admin_client:
        admin_csrf = _login(admin_client, "admin", "1")
        _create_user(admin_client, admin_csrf, login_name="traffic-user")
        user_client = TestClient(app)
        try:
            _login(user_client, "traffic-user", "initial-password")
            assert user_client.get("/api/admin/audit-traffic").status_code == 403
        finally:
            user_client.close()

        with SessionLocal() as db:
            admin_user = db.scalar(select(User).where(User.login_id == "admin@posco.com"))
            assert admin_user is not None
            now = utc_now()
            db.add_all(
                [
                    AuditEvent(
                        organization_id=admin_user.organization_id,
                        actor_user_id=admin_user.id,
                        action=f"traffic_probe_{index}",
                        target_type="test",
                        result="success",
                        created_at=now - timedelta(minutes=2, seconds=index),
                    )
                    for index in range(3)
                ]
            )
            db.add(
                AuditEvent(
                    organization_id=admin_user.organization_id,
                    actor_user_id=admin_user.id,
                    action="outside_traffic_window",
                    target_type="test",
                    result="success",
                    created_at=now - timedelta(minutes=61),
                )
            )
            db.add(
                AuditEvent(
                    organization_id=admin_user.organization_id,
                    actor_user_id=admin_user.id,
                    action="failed_monitoring_probe",
                    target_type="test",
                    result="failure",
                    created_at=now - timedelta(minutes=1),
                )
            )
            db.commit()

        launcher_log = tmp_path / "logs" / "launcher-events.jsonl"
        launcher_log.parent.mkdir(parents=True, exist_ok=True)
        launcher_log.write_text(
            "\n".join(
                [
                    json.dumps(
                        {
                            "timestamp": (now - timedelta(minutes=1)).isoformat(),
                            "event": "automatic_recovery",
                        }
                    ),
                    json.dumps(
                        {
                            "timestamp": (now - timedelta(minutes=2)).isoformat(),
                            "event": "manual_restart",
                        }
                    ),
                    json.dumps(
                        {
                            "timestamp": (now - timedelta(hours=9)).isoformat(),
                            "event": "automatic_recovery",
                        }
                    ),
                    "not-json",
                ]
            ),
            encoding="utf-8",
        )

        response = admin_client.get("/api/admin/audit-traffic?minutes=60")
        assert response.status_code == 200, response.text
        payload = response.json()
        assert payload["timezone"] == "Asia/Seoul"
        assert payload["periodMinutes"] == 60
        assert len(payload["buckets"]) == 60
        assert payload["total"] == sum(bucket["count"] for bucket in payload["buckets"])
        assert payload["peak"] == max(bucket["count"] for bucket in payload["buckets"])
        assert payload["normalTotal"] == sum(
            bucket["normalCount"] for bucket in payload["buckets"]
        )
        assert payload["abnormalTotal"] == sum(
            bucket["abnormalCount"] for bucket in payload["buckets"]
        )
        assert payload["abnormalAuditTotal"] == 1
        assert payload["automaticRecoveryTotal"] == 1
        assert payload["manualRestartTotal"] == 1
        assert payload["abnormalTotal"] == 3
        assert payload["peak"] >= 3
        assert payload["total"] >= 3
        eight_hours = admin_client.get("/api/admin/audit-traffic?minutes=480")
        assert eight_hours.status_code == 200
        assert len(eight_hours.json()["buckets"]) == 480
        assert admin_client.get("/api/admin/audit-traffic?minutes=14").status_code == 422


def test_admin_conversation_view_is_audited(tmp_path: Path) -> None:
    app = _test_app(tmp_path)
    with TestClient(app) as client:
        csrf = _login(client, "admin", "1")
        project_id = client.get("/api/projects").json()[0]["id"]
        conversation = client.post(
            "/api/conversations",
            headers={"X-CSRF-Token": csrf},
            json={"projectId": project_id, "title": "관리자 조회 감사 테스트"},
        )
        assert conversation.status_code == 201
        conversation_id = conversation.json()["id"]

        with SessionLocal() as db:
            message = Message(
                conversation_id=conversation_id,
                role="assistant",
                status="completed",
                canonical_text="=1+1 관리자 의견 조회 대상 답변",
                turn_index=0,
                metadata_json={},
            )
            db.add(message)
            db.commit()
            message_id = message.id

        report = client.post(
            f"/api/messages/{message_id}/reports",
            headers={"X-CSRF-Token": csrf},
            json={"category": "other", "description": "관리 화면에서 확인할 의견"},
        )
        assert report.status_code == 201, report.text

        listing = client.get("/api/admin/conversations?query=조회 감사")
        assert listing.status_code == 200
        assert listing.json()["items"][0]["id"] == conversation_id
        assert listing.json()["items"][0]["feedbackCount"] == 1

        feedback_listing = client.get("/api/admin/conversations?feedback_only=true")
        assert feedback_listing.status_code == 200
        assert [item["id"] for item in feedback_listing.json()["items"]] == [
            conversation_id
        ]

        detail = client.get(f"/api/admin/conversations/{conversation_id}")
        assert detail.status_code == 200
        assert detail.json()["conversation"]["title"] == "관리자 조회 감사 테스트"
        assert (
            detail.json()["feedback"][0]["description"] == "관리 화면에서 확인할 의견"
        )
        assert detail.json()["feedback"][0]["author"]["loginId"] == "admin@posco.com"

        turn_sets = client.get(f"/api/admin/conversations/{conversation_id}/turn-sets")
        assert turn_sets.status_code == 200

        audit = client.get(
            "/api/admin/audit-events"
            f"?action=admin_conversation_viewed&target_id={conversation_id}"
        )
        assert audit.status_code == 200
        assert len(audit.json()["items"]) == 2

        rating = client.put(
            f"/api/messages/{message_id}/rating",
            headers={"X-CSRF-Token": csrf},
            json={"value": "dislike"},
        )
        assert rating.status_code == 200, rating.text

        exported = client.get(
            "/api/admin/conversations/export.xlsx"
            "?query=조회%20감사&feedback_only=true&limit=120"
        )
        assert exported.status_code == 200, exported.text
        assert exported.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "lumina_conversations_" in exported.headers["content-disposition"]

        workbook = load_workbook(BytesIO(exported.content))
        assert workbook.sheetnames == ["대화 분석"]
        analysis_sheet = workbook["대화 분석"]
        headers = {
            cell.value: index for index, cell in enumerate(analysis_sheet[1], start=1)
        }
        assert analysis_sheet.max_row == 2
        assert analysis_sheet.freeze_panes == "A2"
        assert analysis_sheet.auto_filter.ref is not None
        assert (
            analysis_sheet.cell(2, headers["메시지 내용"]).value
            == "=1+1 관리자 의견 조회 대상 답변"
        )
        assert analysis_sheet.cell(2, headers["의견 종류"]).value == "rating + report"
        assert analysis_sheet.cell(2, headers["평가"]).value == "dislike"
        assert analysis_sheet.cell(2, headers["좋아요 수"]).value == 0
        assert analysis_sheet.cell(2, headers["싫어요 수"]).value == 1
        assert analysis_sheet.cell(2, headers["Category"]).value == "other"
        assert analysis_sheet.cell(2, headers["Comment 수"]).value == 1
        assert (
            analysis_sheet.cell(2, headers["Comment"]).value
            == "관리 화면에서 확인할 의견"
        )
        assert (
            analysis_sheet.cell(2, headers["의견 작성자"]).value
            == "admin@posco.com"
        )
        assert analysis_sheet.cell(2, headers["메시지 내용"]).data_type == "s"

        export_audit = client.get(
            "/api/admin/audit-events?action=admin_conversations_exported"
        )
        assert export_audit.status_code == 200
        assert export_audit.json()["items"][0]["metadata"] == {
            "query_used": True,
            "feedback_only": True,
            "limit": 120,
            "conversation_count": 1,
            "message_count": 1,
            "feedback_count": 2,
        }


def test_admin_announcements_are_managed_by_admins_and_visible_to_users(
    tmp_path: Path,
) -> None:
    app = _test_app(tmp_path)
    with TestClient(app) as admin_client:
        admin_csrf = _login(admin_client, "admin", "1")
        _create_user(admin_client, admin_csrf, login_name="announcement-reader")

        user_client = TestClient(app)
        try:
            user_csrf = _login(
                user_client,
                "announcement-reader",
                "initial-password",
            )
            assert user_client.get("/api/admin/announcements").status_code == 403
            assert (
                user_client.post(
                    "/api/admin/announcements",
                    headers={"X-CSRF-Token": user_csrf},
                    json={"title": "권한 없음", "body": "게시할 수 없습니다."},
                ).status_code
                == 403
            )

            created = admin_client.post(
                "/api/admin/announcements",
                headers={"X-CSRF-Token": admin_csrf},
                json={
                    "title": "서비스 점검 안내",
                    "body": "금요일 18시에 점검을 시작합니다.",
                },
            )
            assert created.status_code == 201, created.text
            announcement = created.json()
            assert announcement["author"]["loginId"] == "admin@posco.com"

            for invalid_payload in ({}, {"title": None, "body": None}):
                rejected_update = admin_client.patch(
                    f"/api/admin/announcements/{announcement['id']}",
                    headers={"X-CSRF-Token": admin_csrf},
                    json=invalid_payload,
                )
                assert rejected_update.status_code == 422, rejected_update.text

            user_listing = user_client.get("/api/notifications/announcements")
            assert user_listing.status_code == 200, user_listing.text
            assert user_listing.json()["items"][0]["title"] == "서비스 점검 안내"
            assert user_listing.json()["total"] == 1
            assert user_listing.json()["unreadCount"] == 1
            assert user_listing.json()["items"][0]["readAt"] is None

            marked_read = user_client.post(
                f"/api/notifications/announcements/{announcement['id']}/read",
                headers={"X-CSRF-Token": user_csrf},
            )
            assert marked_read.status_code == 200, marked_read.text
            assert marked_read.json()["readAt"] is not None
            assert user_client.get(
                "/api/notifications/announcements/unread-count"
            ).json() == {"unreadCount": 0}

            updated = admin_client.patch(
                f"/api/admin/announcements/{announcement['id']}",
                headers={"X-CSRF-Token": admin_csrf},
                json={
                    "title": "서비스 점검 시간 변경",
                    "body": "금요일 19시에 점검을 시작합니다.",
                },
            )
            assert updated.status_code == 200, updated.text
            assert updated.json()["title"] == "서비스 점검 시간 변경"
            assert user_client.get(
                "/api/notifications/announcements/unread-count"
            ).json() == {"unreadCount": 1}

            searched = admin_client.get(
                "/api/admin/announcements?query=시간 변경"
            )
            assert searched.status_code == 200
            assert [item["id"] for item in searched.json()["items"]] == [
                announcement["id"]
            ]

            deleted = admin_client.delete(
                f"/api/admin/announcements/{announcement['id']}",
                headers={"X-CSRF-Token": admin_csrf},
            )
            assert deleted.status_code == 204
            assert user_client.get("/api/notifications/announcements").json() == {
                "items": [],
                "total": 0,
                "unreadCount": 0,
            }

        finally:
            user_client.close()

        audit = admin_client.get("/api/admin/audit-events?target_id=" + announcement["id"])
        assert audit.status_code == 200
        assert {item["action"] for item in audit.json()["items"]} == {
            "announcement_created",
            "announcement_read",
            "announcement_updated",
            "announcement_deleted",
        }
